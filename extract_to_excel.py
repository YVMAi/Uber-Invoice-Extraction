#!/usr/bin/env python3
"""
Uber Receipt PDF → Excel Extractor
Parses downloaded Uber receipt PDFs and generates a structured Excel summary.

Usage:
  python extract_to_excel.py --folder ./uber_invoices --output uber_trips_summary.xlsx

Handles two major Uber receipt formats:
  Format A: Older style — full date ("3 April 2025"), 24h times, pipe-separated locations
  Format B: Newer style — abbreviated date ("31 Oct 2025"), 12h times, locations on page 2

Extracts 16 fields per trip including fare breakdown, GST, driver info, and payment status.
"""

import argparse
import re
import sys
from pathlib import Path
from datetime import datetime
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ─────────────────────────────────────────────────────────────────────
# PDF TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────

def extract_text(pdf_path):
    """Read all pages of a PDF and return combined normalized text."""
    try:
        reader = PdfReader(str(pdf_path))
        pages = [(page.extract_text() or "") for page in reader.pages]
        raw = " ".join(pages)
        # Normalize: collapse newlines/extra spaces so regex works across line breaks
        norm = re.sub(r"\n+", " ", raw)
        norm = re.sub(r" {2,}", " ", norm)
        return norm, pages
    except Exception as e:
        return None, str(e)

# ─────────────────────────────────────────────────────────────────────
# FIELD PARSERS — each returns the extracted value or a sensible default
# ─────────────────────────────────────────────────────────────────────

def parse_date(text):
    """
    Uber uses multiple date formats across receipt versions:
      "3 April 2025"   — Format A (full month, day-first)
      "30 Oct 2025"    — Format B (abbreviated month, day-first)
      "Oct 31, 2025"   — Format B variant (US-style, month-first)
      "November 1, 2025" — Format A variant (full month, US-style)
    Returns a datetime or None.
    """
    formats = [
        # Day-first with full month: "3 April 2025"
        (r"(\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4})", "%d %B %Y"),
        # Day-first with abbreviated month: "30 Oct 2025"
        (r"(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4})", "%d %b %Y"),
        # US-style with abbreviated month: "Oct 31, 2025"
        (r"((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s*\d{4})", "%b %d, %Y"),
        # US-style with full month: "November 1, 2025"
        (r"((?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s*\d{4})", "%B %d, %Y"),
    ]
    for pattern, fmt in formats:
        m = re.search(pattern, text)
        if m:
            try:
                return datetime.strptime(m.group(1), fmt)
            except ValueError:
                continue
    return None


def parse_amount(pattern, text):
    """Generic ₹-amount extractor. Returns float or None."""
    m = re.search(pattern, text)
    if m:
        return float(m.group(1).replace(",", ""))
    return None


def parse_total(text):
    """Extract total amount: 'Total ₹608.95' or 'Total ₹ 1,150.56'"""
    return parse_amount(r"Total\s*₹\s*([\d,]+\.?\d*)", text)


def parse_suggested_fare(text):
    """Extract suggested fare (newer receipts only): 'Suggested fare ₹185.53'"""
    return parse_amount(r"Suggested fare\s*₹\s*([\d,]+\.?\d*)", text)


def parse_discount(text):
    """
    Extract promotion/discount amount. Appears as:
      "Promotion -₹107.46"  |  "Rider Promotion -₹3.26"  |  "Promo -₹50.00"
    Returns the discount as a positive float, or 0.0 if none.
    """
    m = re.search(r"(?:Promotion|Promo)\s*-\s*₹\s*([\d,]+\.?\d*)", text)
    return float(m.group(1).replace(",", "")) if m else 0.0


def parse_gst(text):
    """
    Extract GST amount from: "has a GST of ₹59.90 included"
    If "No GST is being recovered" → return "N/A"
    """
    m = re.search(r"GST\s+of\s*₹\s*([\d,]+\.?\d*)", text)
    if m:
        return float(m.group(1).replace(",", ""))
    if re.search(r"No GST is being recovered", text):
        return "N/A"
    return "N/A"


def parse_distance(text):
    """
    Extract distance in km. Formats:
      "19.87 kilometres | 40 min(s)"
      "13.16 kilometres, 22 minutes"
    """
    m = re.search(r"([\d.]+)\s*kilometres", text)
    return float(m.group(1)) if m else None


def parse_duration(text):
    """
    Extract trip duration in minutes. Formats:
      "| 40 min(s)"         →  40
      ", 22 minutes"        →  22
      "| 3 h 54 min(s)"    →  234
      "1 hours 47 minutes"  →  107
    """
    # Hours + minutes: "3 h 54 min(s)" or "1 hours 47 minutes"
    m = re.search(r"(\d+)\s*(?:h|hours?)\s*(\d+)\s*(?:min|minutes?)", text)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    # Minutes only: "| 40 min(s)" or ", 22 minutes"
    m = re.search(r"[|,]\s*(\d+)\s*(?:min|minutes?)", text)
    if m:
        return int(m.group(1))
    return None


def parse_driver(text):
    """
    Extract driver name and rating from:
      "You rode with Vijay 4.80"  →  ("Vijay", 4.80)
      "You rode with RAKESH"      →  ("RAKESH", None)
    Returns (name, rating) tuple.
    """
    # Name + rating: "You rode with Vijay 4.80 Want to review"
    m = re.search(r"You rode with\s+([A-Za-z ]+?)\s+(\d\.\d{1,2})\s", text)
    if m:
        return m.group(1).strip(), float(m.group(2))
    # Name only — stop before "License Plate" / "Fares" / other keywords
    m = re.search(r"You rode with\s+([A-Za-z]+)(?:\s|$)", text)
    if m:
        return m.group(1).strip(), None
    return "", None


def parse_license_plate(text):
    """Extract license plate: 'License Plate: HR55AU8974'"""
    m = re.search(r"License Plate:\s*([A-Z0-9]+)", text)
    return m.group(1).strip() if m else ""


def parse_departure_time(text):
    """
    Extract departure time from the route section.
    Format A: "21:39 | Address"  →  "21:39"
    Format B: After "Trip details...License Plate:..." then "06:44 Address" → "06:44"
    Also handles 12h format: "4:19 pm"
    """
    # Look in route section (after Trip details / distance info)
    after = re.split(r"(?:Trip details|kilometres|minutesLicense|min\(s\))", text)
    if len(after) > 1:
        tail = after[-1]
        m = re.search(r"(\d{1,2}:\d{2}(?:\s*(?:am|pm))?)", tail)
        if m:
            return m.group(1).strip()

    # Fallback: pipe-separated format "21:39 |"
    m = re.search(r"(\d{2}:\d{2})\s*\|", text)
    if m:
        return m.group(1)
    return ""


def parse_locations(text):
    """
    Extract pickup (from) and dropoff (to) addresses.

    Format A (pipe-separated): "21:39 | T3, Multi Level... 22:20 | Inder Mohan..."
    Format B (page 2, no pipe): "06:44 Address1... 07:01 Address2..."

    Addresses end before "You rode with" or "Want to review".
    """
    TIME = r"\d{1,2}:\d{2}(?:\s*(?:am|pm|AM|PM))?"

    # Work on the route section (after Trip details / distance / duration info)
    after = re.split(r"(?:Trip details|kilometres|minutesLicense|min\(s\))", text)

    if len(after) > 1:
        tail = after[-1]
        matches = re.findall(
            TIME + r"\s*\|?\s*([A-Z0-9].+?)(?=\s*" + TIME + r"\s|You rode with|Want to review|$)",
            tail,
        )
        if len(matches) >= 2:
            return matches[0].strip().rstrip(","), matches[1].strip().rstrip(",")
        if len(matches) == 1:
            return matches[0].strip().rstrip(","), ""

    # Fallback: pipe-separated format from full text
    end = r"(?=\s*\d{1,2}:\d{2}[\s|]|(?:Uber|Go Rentals|Premier)\b|\d+\.?\d*\s*kilometres|You rode with|Fares are)"
    matches = re.findall(r"\d{1,2}:\d{2}(?:\s*(?:am|pm))?\s*\|\s*(.+?)" + end, text)
    if len(matches) >= 2:
        return matches[0].strip().rstrip(","), matches[1].strip().rstrip(",")
    if len(matches) == 1:
        return matches[0].strip().rstrip(","), ""

    return "", ""


def parse_payments(text):
    """
    Extract payment method and status from the Payments section.

    Uber receipts may list multiple payment attempts:
      "Cash ₹141.44"                          →  Cash, Success
      "UPI Scan and Pay ₹141.44 ... Failed"   →  UPI failed, check if Cash succeeded
      "Uber Cash -₹19.25"                     →  Partial credit (not primary method)

    Returns (payment_method, payment_status).
    """
    # Find the Payments section
    payments_section = re.split(r"Payments", text, maxsplit=1)
    if len(payments_section) < 2:
        # Format A: payment method at very start of text ("Cash 03/04/2025...")
        top = text[:150]
        for method in ["Cash", "UPI Scan and Pay", "UPI", "Card", "Visa", "Paytm", "Google Pay", "PhonePe"]:
            if method in top:
                return method, "Success"
        return "", "Unknown"

    pay_text = payments_section[1][:500]

    # Match payment lines with optional "Failed" flag
    payment_lines = re.findall(
        r"(Cash|UPI Scan and Pay|UPI|Uber Cash|Paytm|Card|Visa|Mastercard|RuPay|Debit Card|Credit Card|Google Pay|PhonePe)"
        r"\s*[-]?₹\s*[\d,]+\.?\d*"
        r"(?:\s*\d{1,2}/\d{1,2}/\d{2,4}\s*\d{1,2}:\d{2}(?:\s*(?:am|pm))?)?"
        r"(\s*Failed)?",
        pay_text,
        re.IGNORECASE,
    )

    if not payment_lines:
        # Check payment section and start of text for method keywords
        for method in ["Cash", "UPI Scan and Pay", "UPI", "Card", "Visa", "Paytm", "Google Pay", "PhonePe", "Uber Cash"]:
            if method in pay_text:
                return method, "Success"
        top = text[:100] if len(text) > 100 else text
        for method in ["Cash", "UPI", "Card", "Visa", "Paytm", "Google Pay", "PhonePe"]:
            if method in top:
                return method, "Success"
        return "", "Unknown"

    # Classify each payment line as successful or failed
    successful = []
    failed = []
    for method, fail_flag in payment_lines:
        if fail_flag and "Failed" in fail_flag:
            failed.append(method)
        elif method == "Uber Cash":
            continue  # Uber Cash is a partial credit, not primary method
        else:
            successful.append(method)

    if successful:
        return successful[0], "Success"
    elif failed:
        return failed[0], "Failed"
    else:
        return "Uber Cash", "Success"


# ─────────────────────────────────────────────────────────────────────
# MAIN EXTRACTION PIPELINE
# ─────────────────────────────────────────────────────────────────────

def extract_receipt(pdf_path):
    """Parse one Uber receipt PDF and return a dict of all extracted fields."""
    text, pages = extract_text(pdf_path)
    if text is None:
        return None, f"Read error: {pages}"

    if not text.strip():
        return None, "Empty PDF"

    trip_date = parse_date(text)
    total = parse_total(text)
    suggested = parse_suggested_fare(text)
    discount = parse_discount(text)
    gst = parse_gst(text)
    distance = parse_distance(text)
    duration = parse_duration(text)
    driver_name, driver_rating = parse_driver(text)
    plate = parse_license_plate(text)
    dep_time = parse_departure_time(text)
    from_addr, to_addr = parse_locations(text)
    pay_method, pay_status = parse_payments(text)

    return {
        "file_name": pdf_path.name,
        "date": trip_date,
        "departure_time": dep_time,
        "driver": driver_name,
        "license_plate": plate,
        "from": from_addr,
        "to": to_addr,
        "distance_km": distance,
        "duration_min": duration,
        "suggested_fare": suggested if suggested else total,
        "discount": discount,
        "amount_paid": total,
        "gst": gst,
        "payment_method": pay_method,
        "payment_status": pay_status,
        "driver_rating": driver_rating,
    }, None


# ─────────────────────────────────────────────────────────────────────
# DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────

def deduplicate(records):
    """Remove duplicate trips based on (date + amount + license_plate)."""
    seen = set()
    unique = []
    dupes = 0
    for r in records:
        key = (
            r["date"].strftime("%Y-%m-%d") if r["date"] else "",
            r["amount_paid"],
            r["license_plate"],
        )
        if key in seen:
            dupes += 1
            continue
        seen.add(key)
        unique.append(r)
    return unique, dupes


# ─────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────

COLUMNS = [
    ("File Name", "file_name", 42),
    ("Date", "date", 14),
    ("Departure Time", "departure_time", 15),
    ("Driver", "driver", 18),
    ("License Plate", "license_plate", 16),
    ("From", "from", 55),
    ("To", "to", 55),
    ("Distance (km)", "distance_km", 14),
    ("Duration (min)", "duration_min", 14),
    ("Suggested Fare (₹)", "suggested_fare", 18),
    ("Discount (₹)", "discount", 14),
    ("Amount Paid (₹)", "amount_paid", 18),
    ("GST", "gst", 12),
    ("Payment Method", "payment_method", 18),
    ("Payment Status", "payment_status", 16),
    ("Driver Rating", "driver_rating", 14),
]


def write_excel(records, output_path):
    """Generate a formatted Excel workbook with all trip data and totals."""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Rides"

    # Styles
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="333333")
    hdr_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=10)
    border = Border(
        left=Side("thin", "DDDDDD"), right=Side("thin", "DDDDDD"),
        top=Side("thin", "DDDDDD"), bottom=Side("thin", "DDDDDD"),
    )

    # Header row (bold, white text on dark background)
    for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
        ws.column_dimensions[c.column_letter].width = width

    ws.freeze_panes = "A2"  # Freeze header row

    # Data rows
    for row_idx, rec in enumerate(records, 2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val = rec.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border

            if key == "date" and isinstance(val, datetime):
                cell.number_format = "DD/MM/YYYY"
            elif key in ("suggested_fare", "discount", "amount_paid") and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
            elif key == "gst" and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
            elif key == "distance_km" and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"

    # Summary / totals row with SUM formulas
    last = len(records) + 1
    total_row = last + 1
    bold = Font(name="Arial", bold=True, size=11)
    grey_fill = PatternFill("solid", fgColor="F2F2F2")

    ws.cell(row=total_row, column=1, value="TOTAL").font = bold

    # Distance (col H), Suggested Fare (col J), Discount (col K), Amount Paid (col L)
    for col, letter in [(8, "H"), (10, "J"), (11, "K"), (12, "L")]:
        ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{last})").font = bold
        ws.cell(row=total_row, column=col).number_format = "#,##0.00"

    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).fill = grey_fill

    wb.save(str(output_path))


# ─────────────────────────────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Extract Uber receipt PDFs to Excel")
    parser.add_argument("--folder", required=True, help="Folder containing receipt PDFs (searched recursively)")
    parser.add_argument("--output", default="uber_trips_summary.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists():
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    pdf_files = sorted(folder.rglob("*.pdf"))
    if not pdf_files:
        print(f"No PDF files found in {folder}")
        sys.exit(1)

    print(f"Found {len(pdf_files)} PDF files in {folder}\n")

    records = []
    failures = []

    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"  [{i:3d}/{len(pdf_files)}] {pdf_path.name}", end="")
        data, error = extract_receipt(pdf_path)
        if error:
            print(f"  ⚠ {error}")
            failures.append((pdf_path.name, error))
            continue

        missing = []
        if data["date"] is None: missing.append("date")
        if data["amount_paid"] is None: missing.append("total")
        if not data["from"]: missing.append("from")
        if missing:
            print(f"  ⚠ missing: {', '.join(missing)}", end="")

        records.append(data)
        print()

    # Deduplicate
    records, dupes = deduplicate(records)
    if dupes:
        print(f"\n  Removed {dupes} duplicate(s)")

    # Sort by date
    records.sort(key=lambda r: r["date"] or datetime.min)

    # Write Excel
    output = Path(args.output)
    write_excel(records, output)

    # Console summary
    total_amount = sum(r["amount_paid"] or 0 for r in records)
    total_distance = sum(r["distance_km"] or 0 for r in records)

    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"  Total trips     : {len(records)}")
    print(f"  Total amount    : ₹{total_amount:,.2f}")
    print(f"  Total distance  : {total_distance:,.1f} km")
    print(f"  Duplicates      : {dupes}")
    print(f"  Failed to parse : {len(failures)}")
    if failures:
        for fname, err in failures[:10]:
            print(f"    - {fname}: {err}")
    print(f"  Output file     : {output.resolve()}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
